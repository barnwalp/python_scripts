import pandas as pd
from openpyxl import load_workbook
from datetime import date, timedelta
import numpy as np
import openpyxl


def cell_to_num(char):
    char = char.upper()
    if char.isalpha():
        num = ord(char) - 64
    else:
        num = 0
    return num


print(cell_to_num('a'))

# Get the file from the downloads folder
first_file = '/mnt/c/Users/panka/Downloads/Nozzle Sales Report.xlsx'
second_file = '/mnt/c/Users/panka/Downloads/Nozzle Sales Report (1).xlsx'


def data_cleaning(path):
    wb = load_workbook(path)
    ws = wb.active
    # Unmerge cell
    ws.unmerge_cells('O1:P1')
    ws.unmerge_cells('Q1:R1')
    ws.unmerge_cells('S1:T1')
    ws.unmerge_cells('U1:V1')
    if path[-7] == '1':
        # renaming issue_sale value with dates
        ws['S1'] = str(date.today() - timedelta(days=3))
        ws['T1'] = str(date.today() - timedelta(days=4))
    else:
        # renaming issue_sale value with dates
        ws['S1'] = str(date.today() - timedelta(days=1))
        ws['T1'] = str(date.today() - timedelta(days=2))

    print(ws['S1'].value)
    print(ws['T1'].value)
    # Delete column U:V
    ws.delete_cols(cell_to_num('U'), cell_to_num('V')-cell_to_num('U')+1)
    # Delete column O:R
    # print(cell_to_num('R')-cell_to_num('O')+1)
    ws.delete_cols(cell_to_num('O'), cell_to_num('R')-cell_to_num('O')+1)
    # Delete column K:M
    ws.delete_cols(cell_to_num('K'), cell_to_num('M')-cell_to_num('K')+1)
    # Delete column B:F
    ws.delete_cols(cell_to_num('B'), cell_to_num('F')-cell_to_num('B')+1)
    # Delete 2nd row
    ws.delete_rows(2)
    wb.save(path)


def create_pivot(path):
    # reading excel file using pandas
    old_df = pd.read_excel(first_file)
    # print(old_df.columns)

    # filter the table using product value
    df = old_df[
        (old_df['Product'] == 'XP') |
        (old_df['Product'] == 'HS') |
        (old_df['Product'] == 'MS')
    ]

    df.loc[df['Product'] == 'XP', 'Product'] = 'MS'
    print(df.columns)
    # wb_first = load_workbook(first_file)
    # ws = wb_first.active
    # print(ws['G1'].value)

    pvt_all = pd.pivot_table(df,
                             index=["SA"],
                             columns=['Product'],
                             values=["2020-07-04", "2020-07-03"],
                             aggfunc=np.sum)

    return pvt_all


# data_cleaning(first_file)
# data_cleaning(second_file)

# print(type(create_pivot(first_file)))
# print('------------------------------------------------------------')
# print(create_pivot(second_file))

# concatenate two pivot table into single dataframe
pvt = pd.concat([create_pivot(first_file), create_pivot(second_file)], axis=1)


# Writing panda pivot table to excel sheet
with pd.ExcelWriter(first_file, engine='openpyxl') as writer:
    writer.book = openpyxl.load_workbook(first_file)
    pvt.to_excel(writer, "pivot sheet", index=True)
