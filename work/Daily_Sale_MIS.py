#!/usr/bin/env python
# coding: utf-8

# In[1]:


from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import os
from datetime import date, timedelta
import time


# In[2]:


# Deleting existing .xlsx files in Download folders
# -------------------------------------------------

# Download folder path
path = "C:/Users/panka/Downloads/"

# Getting all the files in a variable
files_in_downloads = os.listdir(path)

# Getting all the xlsx files in a variable
excel_files = [file for file in files_in_downloads if file.endswith('xlsx')]

# Deleting all .xlsx files after a simple availability check
if len(excel_files) > 0:
    for file in excel_files:
        path_to_file = os.path.join(path, file)
        os.remove(path_to_file)
        print(f'{file} has been deleted')


# In[3]:


browser = webdriver.Chrome()
browser.get('https://spandan.indianoil.co.in/RAPHASE6/SigninPage')
password = os.getenv('CDRSM_PASS')
user = os.getenv('CDRSM_UNAME')


# In[4]:


# Changing the dropdown in the select form
# select = Select(browser.find_element_by_css_selector("#type.form-control"))
# select.select_by_visible_text("Dealer")

browser.find_element_by_name('username').send_keys(user)
browser.find_element_by_name('password').send_keys(password)


# In[5]:


# the trick to avoid captcha is to wait for user to enter it manually
# and click the submit button, then wait until an element from the
# next page is located i.e. 'global_ro_code' in this case
element = WebDriverWait(browser, 35).until(
    EC.presence_of_element_located((By.ID, 'global_ro_code'))
)


# In[6]:


# data = '152110-AVTAR FILLING STATION(Sambalpur DO)'
# browser.find_element_by_name('global_ro_code').send_keys(data)
"""
menu_1 = browser.find_element_by_css_selector(".navbar-nav#tay .dropdown")
menu_1.click()
menu_2 = browser.find_element_by_css_selector(".navbar-nav#tay .dropdown .dropdown-menu>li")
menu_2.click()
"""

# Get into the sub-menus of the navigation bar
browser.find_element_by_link_text('View Reports').click()
browser.find_element_by_link_text('Sales Reports').click()
browser.find_element_by_link_text('Nozzle Sales Report').click()

# move mouse curser by xoffset and yoffset to unselect the visible menu-bar
webdriver.ActionChains(browser).move_by_offset(10, 10).perform()


# In[7]:


browser.implicitly_wait(10)

# selecting the dropdown in the state office input
select = Select(browser.find_element_by_id("SO"))
select.select_by_visible_text("Odisha State Office")

# selecting the dropdown in the divisional office input
select = Select(browser.find_element_by_id("DO"))
select.select_by_visible_text("Sambalpur DO")

# selecting the dropdown in the sales area input
select = Select(browser.find_element_by_id("SA"))

browser.implicitly_wait(10)

# since the values in the dropdown activates after selection of preceding input.( eg
# SA dropdown activates only after Divisional office get selected); hence it is
# prudent to wait untill all the text is present in the box

element = WebDriverWait(browser, 20).until(
    EC.text_to_be_present_in_element((By.ID, 'SA'), 'ALL')
)

# Download the first excel file
browser.find_element_by_id('excelButton').click()


# In[8]:


# Check if the first file is downloaded
path_to_file = 'C:/Users/panka/Downloads/Nozzle Sales Report.xlsx'
while not os.path.exists(path_to_file):
    time.sleep(2)

print(f'{path_to_file} exists')


# In[9]:


# selecting date using datepicker calendar
datepicker = browser.find_element_by_id('datetimepicker')
# opening the calendar using the datepicker
datepicker.click()

browser.implicitly_wait(10)

# getting today's day
current_day = date.today().day

# selecting all elements with the class 'ui-state-defualt'. this
# element will only appear after the calendar is open through datepicker
# days is a list of selenium specific data structure
days = browser.find_elements_by_css_selector('.day')

print(current_day)

# this will print the text of the 18th value of the days list
print(days[17].text)
# days[17].click()
# print(days)
# print(days[6])
# print(type(days[6]))

# Since there are double occurence of few tail days such as 29, 30, 31
# in the calendar, it is imperative that correct day is choosen while
# selecting the date. It can be inferred that current date will always
# come after day 1, so to locate the current day, one need to traverse
# from day 1 to day2 ....... day 31 and discard old days of previous month

# Now one need to find the index of the day 1 of the current month
for index, val in enumerate(days):
    if val.text == '1':
        day_1_index = index
        print(day_1_index)
        break

# Given the index of day 1, we have to first find the index of today
# in the calendar. Thereafter one can navigate to the preceeding days
# using the index of the days list. Following will find the index of
# current date and will store it in anchor variable.

for index, val in enumerate(days):
    if index >= day_1_index:
        if val.text == str(current_day):
            print(val.text)
            anchor = index
            print(anchor)
            break

# This will select the day, 4 days prior from the current day.
days[(anchor-4)].click()

browser.implicitly_wait(10)

# selecting second date calendar using datepicker1
# ----------------------------------------
datepicker_2 = browser.find_element_by_id('datetimepicker1')
datepicker_2.click()

browser.implicitly_wait(10)

days = browser.find_elements_by_css_selector('.day')

# This will select the day, 3 days prior from the current day.
days[(anchor-3)].click()


# In[10]:


# print(select.options[0])
#print(len(select.options))

# Download the excel file
browser.find_element_by_id('excelButton').click()


# In[11]:


import pandas as pd
from openpyxl import load_workbook
import openpyxl
import numpy as np
from datetime import date, timedelta


# In[12]:


def cell_to_num(char):
    char = char.upper()
    if char.isalpha():
        num = ord(char) - 64
    else:
        num = 0
    return num
print(cell_to_num('a'))


# In[13]:


# Windows file system
first_file = 'C:/Users/panka/Downloads/Nozzle Sales Report.xlsx'
second_file = 'C:/Users/panka/Downloads/Nozzle Sales Report (1).xlsx'
"""
# linux file system
first_file = '/mnt/c/Users/panka/Downloads/Nozzle Sales Report.xlsx'
second_file = '/mnt/c/Users/panka/Downloads/Nozzle Sales Report (1).xlsx'
"""
wb_check = load_workbook(first_file)


# In[14]:


def data_cleaning(path):
    wb = load_workbook(path)
    ws = wb.active

    #Unmerge cell
    ws.unmerge_cells('O1:P1')
    ws.unmerge_cells('Q1:R1')
    ws.unmerge_cells('S1:T1')
    ws.unmerge_cells('U1:V1')
    if path[-7] == '1':
        #renaming issue_sale value with dates
        ws['S1'] = str(date.today() - timedelta(days=4))
        ws['T1'] = str(date.today() - timedelta(days=3))
    else:
        ws['S1'] = str(date.today() - timedelta(days=2))
        ws['T1'] = str(date.today() - timedelta(days=1))

    print(ws['S1'].value)
    print(ws['T1'].value)

    #Delete column U:V
    ws.delete_cols(cell_to_num('U'), cell_to_num('V')-cell_to_num('U')+1)
    #Delete column O:R
    ws.delete_cols(cell_to_num('O'), cell_to_num('R')-cell_to_num('O')+1)
    #Delete column K:M
    ws.delete_cols(cell_to_num('K'), cell_to_num('M')-cell_to_num('K')+1)
    #Delete column B:F
    ws.delete_cols(cell_to_num('B'), cell_to_num('F')-cell_to_num('B')+1)
    #Delete second row
    ws.delete_rows(2)
    wb.save(path)


# In[15]:


def create_pivot(path):
    #reading excel file using pandas
    old_df = pd.read_excel(path)
    current_wb = load_workbook(path)
    ws = current_wb.active

    #filter out the 'NA' value in the product column
    df = old_df[
        (old_df['Product'] == 'XP') |
        (old_df['Product'] == 'HS') |
        (old_df['Product'] == 'MS')
    ]

    df.loc[df['Product'] == 'XP', 'Product'] = 'MS'

    pvt_all = pd.pivot_table(df,
                            index=["SA"],
                            columns=['Product'],
                            values=[ws['H1'].value, ws['G1'].value],
                            aggfunc=np.sum)
    return pvt_all


# In[16]:


data_cleaning(first_file)
data_cleaning(second_file)


# In[17]:


pvt = pd.concat([create_pivot(first_file), create_pivot(second_file)], axis=1)
print(pvt)


# In[18]:


with pd.ExcelWriter(first_file, engine='openpyxl') as writer:
    writer.book = load_workbook(first_file)
    pvt.to_excel(writer, 'pivot sheet', index=True)


# In[19]:


df = pd.read_excel(first_file, sheet_name='pivot sheet')
print(df.columns.values)


# In[20]:


cols = df.columns.values
order = [0, 3, 4, 1, 2, 7, 8, 5, 6]
cols = [cols[i] for i in order]
print(cols)


# In[21]:


df = df[cols]
print(df)


# In[22]:


with pd.ExcelWriter(first_file, engine='openpyxl') as writer:
    writer.book = load_workbook(first_file)
    df.to_excel(writer, 'summary', index=False)


# In[23]:


wb = load_workbook(first_file)
ws = wb['summary']
ws.delete_rows(3)
check = 0
for value in ws.iter_rows(min_row=1, max_row=1):
    for cell in value:
        if check % 2 == 0:
            cell.value = ""
        check += 1
        print(cell.value)
# print(df.columns.values)
print(ws['A3'].value)


# In[24]:


from exchangelib import Account, Message, Credentials, HTMLBody
from exchangelib import Configuration, DELEGATE
import os


# In[25]:


outlook_user = os.environ.get('OUTLOOK_USER')
outlook_password = os.environ.get('OUTLOOK_PASS')
outlook_server = os.environ.get('OUTLOOK_SERVER')
outlook_email = os.environ.get('OUTLOOK_EMAIL')

print(os.getenv('OUTLOOK_USER'))
print(os.getenv('OUTLOOK_PASS'))


# In[26]:


credentials = Credentials(username=outlook_user,
                         password=outlook_password
                         )
config = Configuration(server=outlook_server,
                      credentials=credentials)
account = Account(primary_smtp_address=outlook_email,
                 config=config,
                 autodiscover=False,
                 access_type=DELEGATE)


# In[27]:


def create_html(ws):
    html = """
    <html><body><h3>Nozzle Sales Report:</h3>
    <table border=1>
    """
    # entering header data in html
    for whole_row in ws.iter_rows(min_row=1, max_row=2):
        html += "<tr>"
        for row in whole_row:
            html += f'<th>{row.value}</th>'
        html += "</tr>"
    # entering table data in html
    for whole_row in ws.iter_rows(min_row=3):
        html += "<tr>"
        for row in whole_row:
            if isinstance(row.value, float):
                html += f'<td>{round(row.value, 2)}</td>'
                print(round(row.value, 2))
            else:
                html += f'<td>{row.value}</td>'
        html += "</tr>"
    html += """
    </table>
    <p>
    <br>
    With Regards<br>
    Pankaj Barnwal<br>
    </p>
    </body>
    </html>
    """
    return html


# In[28]:


msg = Message(
    account=account,
    subject="Nozzle Sales Report - Sambalpur DO",
    body=HTMLBody(create_html(ws)),
    to_recipients=['barnwalp@indianoil.in']
)

msg.send_and_save()


# In[ ]:




