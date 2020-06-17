from openpyxl import load_workbook
from datetime import datetime, date, time, timezone


wb = load_workbook(filename='automation mis.xlsx')

# get the worksheet to work on
ws1 = wb['Online RO list']

# Getting a value of a cell
print(f'Value of D12 cell of first worksheet is : {ws1["D12"].value}')

"""
for char in 'ABCDEFGH':
    str = char + "1"
    print(ws1[str].value)
"""
ws4 = wb.create_sheet("testing")

for ws in wb:
    print(ws.title)

# python datetime module
print(f'current date and time: {datetime.now()}')
print(f'Date is: {date(2005, 7, 14)}')
print(f'Time is: {time(12, 30)}')
print(f'datetime is: {datetime.combine(date(2005, 7, 14), time(12, 30))}')
print(f'datetime with timezone: {datetime.now(timezone.utc)}')
print(f'datetime: {datetime.strptime("17/06/20 08:30", "%d/%m/%y %H:%M")}')
