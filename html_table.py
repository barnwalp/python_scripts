import os
from openpyxl import load_workbook


wb = load_workbook(filename="eRACTS Complaint.xlsx")
ws = wb["mail"]

# traverse cells in a spreadsheet column
# for cell in ws['B']:
    # print(cell.value)

for whole_row in ws.iter_rows(min_row=1, max_row=1):
    for row in whole_row:
        print(row.value)

def wrap(a, tag):
    "Wraps in <td> tag the a"
    tag1 = tag
    if tag == "table":
        tag1 = "table border=1"
    if tag == "td" and a.strip().replace(".", "").isdigit():
        print(a, "is a number")
        tag1 = "td style=\"text-align:right\""
    return f"<{tag1}>{a}</{tag}>"


def split(tab):
    "Splits a multiline string in a list of items divided by comma for line"
    tab = tab.splitlines()
    for n, row in enumerate(tab):
        tab[n] = row.split(",")
    return tab


def table(tab):
    html = ''  # contain HTML
    for n, x in enumerate(tab):
        print(f'line is: {x}')
        for a in x:
            print(f'word is: {a}')
            html += "<tr>"
    html = wrap(html, "table")
    return html


data = """
Impiegato, Performance, data
Rossi Mario, 1000, 1/2/2018
Baldo Franco, 2000, 1/2/2018
"""
data = data[1:-1]
print(data)

data = table(data)

with open("Simple.html", "w", encoding="utf-8") as filehtml:
    filehtml.write(data)

os.system("Simple.html")
