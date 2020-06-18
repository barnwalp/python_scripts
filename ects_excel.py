from openpyxl import load_workbook


wb = load_workbook(filename='eRACTS Complaint.xlsx')
ws = wb.active

# ws.unmerge_cells('A1:AI1')
# ws.unmerge_cells('A2:AI2')
# ws.delete_rows(1, 2)

# ws.delete_cols(21, 13)
# ws.delete_cols(12, 3)

# ws.insert_cols(19)
# ws['S1'].value = "Detailed complaint"


# creating a function for filter to filter out None and int types
def is_not_str(x):
    # isinstance(x, int) will check if x is integer
    if x is None or isinstance(x, int):
        return False
    else:
        return True


"""
for cell in ws['S']:
    value_1 = f'P{cell.row}'
    value_2 = f'Q{cell.row}'
    value_3 = f'R{cell.row}'
    merge_list = [ws[value_1].value,
                  ws[value_2].value,
                  ws[value_3].value
                  ]
    # filter will return only those value which are neither None nor int
    cell.value = ', '.join(filter(is_not_str, merge_list))
"""

# copy entire worksheet for mail
# ws2 = wb.copy_worksheet(ws)
# ws2.title = "mail"

ws2 = wb["mail"]
# ws2.delete_cols(13, 6)
# ws2.delete_cols(8, 4)
# ws2.delete_cols(4, 2)
wb.save('eRACTS Complaint.xlsx')
