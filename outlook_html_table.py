import os
from openpyxl import load_workbook
from exchangelib import Account, Message, Credentials, HTMLBody
from exchangelib import Configuration, DELEGATE

# Getting the sensitive information from environment variables
outlook_user = os.environ.get('OUTLOOK_USER')
outlook_password = os.environ.get('OUTLOOK_PASS')
outlook_server = os.environ.get('OUTLOOK_SERVER')
outlook_email = os.environ.get('OUTLOOK_EMAIL')

# Using necessary credential and config to connect exchange server
credentials = Credentials(username=outlook_user, password=outlook_password)
config = Configuration(server=outlook_server,
                       credentials=credentials)
account = Account(primary_smtp_address=outlook_email,
                  config=config, autodiscover=False,
                  access_type=DELEGATE)

wb = load_workbook(filename="eRACTS Complaints 230620.xlsx")


def create_mail_table(wb):
    ws = wb.active
    ws.unmerge_cells('A1:AI1')
    ws.unmerge_cells('A2:AI2')
    ws.delete_rows(1, 2)

    ws.delete_cols(21, 13)
    ws.delete_cols(12, 3)

    ws.insert_cols(19)
    ws['S1'].value = "Detailed complaint"

    # creating a function for filter to filter out None and int types

    def is_not_str(x):
        # isinstance(x, int) will check if x is integer
        if x is None or isinstance(x, int):
            return False
        else:
            return True

    for cell in ws['S']:
        value_1 = f'P{cell.row}'
        value_2 = f'Q{cell.row}'
        value_3 = f'R{cell.row}'
        merge_list = [ws[value_1].value,
                      ws[value_2].value,
                      ws[value_3].value
                      ]
    # filter will return only those value which are neither None nor int
    cell.value = ', '.join(filter(is_not_str, merge_list))

    # copy entire worksheet for mail
    ws2 = wb.copy_worksheet(ws)
    ws2.title = "mail"

    ws2 = wb["mail"]
    ws2.delete_cols(13, 6)
    ws2.delete_cols(8, 4)
    ws2.delete_cols(4, 2)

    filter_value = {'Auto-Closed for Dealer Late Acknowledgement',
                    'Service Completed',
                    'Auto-Closed on Alarm Normalization'
                    }
    # Deleting all complaints which have been closed
    for cell in ws2['H']:
        if cell.row > 1:
            if cell.value in filter_value:
                print(cell.value)
                ws2.delete_rows(cell.row)
    ws2.delete_cols(8)
    ws2['H1'] = "Current Status of complaint"
    ws2['I1'] = "Target Date"

    wb.save('eRACTS Complaint.xlsx')
    return wb


wb = create_mail_table(wb)


def create_html(wb):
    ws = wb["mail"]

    # contains content in HTML format
    html = """
    <html><body><h3>Pl find the current list of eRACTS complaint which have not yet been resolved, pl go through each case and provide current status of complaint and target date of rectification:</h3>
    """

    for whole_row in ws.iter_rows(min_row=1, max_row=1):
        html += "<table border=1><tr>"
        for row in whole_row:
            html += f'<th>{row.value}</th>'
        html += "</tr>"

    for whole_row in ws.iter_rows(min_row=2):
        html += "<tr>"
        for row in whole_row:
            if row.value is not None:
                html += f'<td>{row.value}</td>'
            else:
                html += f'<td>{" "}</td>'
        html += "</tr>"

    html += """
    </table>
    <p>
    <br>
    With Regards<br>
    Pankaj Barnwal<br>
    </p>
    </body>
    </html>
    """
    # print(html)
    return html


msg = Message(
    account=account,
    subject="eRACTS complaint status - 230620 - Sambalpur DO",
    body=HTMLBody(create_html(wb)),
    to_recipients=['barnwalp@indianoil.in']
)

msg.send_and_save()

"""
with open("Simple.html", "w", encoding="utf-8") as filehtml:
    filehtml.write(html)

os.system("Simple.html")
"""
