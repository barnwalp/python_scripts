import pandas as pd
from openpyxl import load_workbook
from datetime import date, timedelta


def cell_to_num(char):
    char = char.upper()
    if char.isalpha():
        num = ord(char) - 64
    else:
        num = 0
    return num


print(cell_to_num('a'))

# Get the file from the downloads folder
path = '/mnt/c/Users/panka/Downloads/Nozzle Sale Report.xlsx'
wb = load_workbook(path)
ws = wb.active

# Unmerge cell S1:T1
ws.unmerge_cells('S1:T1')
# renaming issue value with dates
ws['S1'] = date.today() - timedelta(days=1)
ws['T1'] = date.today() - timedelta(days=2)
print(ws['S1'].value)
print(ws['T1'].value)
# Delete column U:V
ws.delete_cols(cell_to_num('U'), cell_to_num('V')-cell_to_num('U')+1)
# Delete column O:R
# print(cell_to_num('R')-cell_to_num('O')+1)
ws.delete_cols(cell_to_num('O'), cell_to_num('R')-cell_to_num('O')+1)
# Delete column K:M
ws.delete_cols(cell_to_num('K'), cell_to_num('M')-cell_to_num('K')+1)
# Delete column B:F
ws.delete_cols(cell_to_num('B'), cell_to_num('F')-cell_to_num('B')+1)
# Delete 2nd row
ws.delete_rows(2)

wb.save(path)
